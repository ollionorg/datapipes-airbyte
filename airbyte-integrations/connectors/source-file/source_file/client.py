#
# Copyright (c) 2023 Airbyte, Inc., all rights reserved.
#


import json
import logging
import os
import sys
import re
import tempfile
import traceback
import urllib
from os import environ, getcwd
from typing import Iterable, Tuple
from urllib.parse import urlparse
from zipfile import BadZipFile

import backoff
import boto3
import botocore
import google
import numpy as np
import pandas as pd
import pyarrow.parquet as pq
import smart_open
import smart_open.ssh
from airbyte_cdk.entrypoint import logger
from airbyte_cdk.models import AirbyteStream, FailureType, SyncMode
from airbyte_cdk.utils import AirbyteTracedException, is_cloud_environment
from azure.storage.blob import BlobServiceClient
from genson import SchemaBuilder
from google.cloud.storage import Client as GCSClient
from google.oauth2 import service_account
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pandas.errors import ParserError
from paramiko import SSHException
from smart_open import open as s_open
from urllib3.exceptions import ProtocolError
from pandas.api.types import is_datetime64_any_dtype as is_datetime
from pandas.api.types import is_timedelta64_dtype as is_timedelta
from yaml import safe_load
from .encryption_client.pgp import Pgp
from .utils import LOCAL_STORAGE_NAME, backoff_handler

SSH_TIMEOUT = 60

# Force the log level of the smart-open logger to ERROR - https://github.com/airbytehq/airbyte/pull/27157
logging.getLogger("smart_open").setLevel(logging.ERROR)


class ConfigurationError(Exception):
    """Client mis-configured"""


class PermissionsError(Exception):
    """User don't have enough permissions"""


class URLFile:
    """Class to manage read from file located at different providers

    Supported examples of URL this class can accept are as follows:
    ```
        s3://my_bucket/my_key
        s3://my_key:my_secret@my_bucket/my_key
        gs://my_bucket/my_blob
        hdfs:///path/file (not tested)
        hdfs://path/file (not tested)
        webhdfs://host:port/path/file (not tested)
        ./local/path/file
        ~/local/path/file
        local/path/file
        ./local/path/file.gz
        file:///home/user/file
        file:///home/user/file.bz2
        [ssh|scp|sftp]://username@host//path/file
        [ssh|scp|sftp]://username@host/path/file
        [ssh|scp|sftp]://username:password@host/path/file
    ```
    """

    def __init__(self, url: str, provider: dict, binary=None, encoding=None):
        self._url = url
        self._provider = provider
        self._file = None
        self.args = {
            "mode": "rb" if binary else "r",
            "encoding": encoding,
        }

    def __enter__(self):
        return self._file

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    @property
    def full_url(self):
        return f"{self.storage_scheme}{self.url}"

    def close(self):
        if self._file:
            self._file.close()
            self._file = None

    def backoff_giveup(self, error):
        # https://github.com/airbytehq/oncall/issues/1954
        if isinstance(error, SSHException) and str(error).startswith("Error reading SSH protocol banner"):
            # We need to clear smart_open internal _SSH cache from the previous attempt, otherwise:
            # SSHException('SSH session not active')
            # will be raised
            smart_open.ssh._SSH.clear()
            return False
        return True

    def open(self):
        self.close()
        _open = backoff.on_exception(backoff.expo, Exception, max_tries=5, giveup=self.backoff_giveup)(self._open)
        try:
            self._file = _open()
        except google.api_core.exceptions.NotFound as err:
            raise FileNotFoundError(self.url) from err
        return self

    def create_temp_file(self, prefix: str, content: str) -> str:
        private_ssh_key_path = f"{getcwd()}/{prefix}"
        with open(private_ssh_key_path, "w") as t:
            t.write(content)
        return private_ssh_key_path

    def _open(self):
        storage = self.storage_scheme
        url = self.url

        if storage == "gs://":
            return self._open_gcs_url()
        elif storage == "s3://":
            return self._open_aws_url()
        elif storage == "azure://":
            return self._open_azblob_url()
        elif storage == "webhdfs://":
            host = self._provider["host"]
            port = self._provider["port"]
            return smart_open.open(f"webhdfs://{host}:{port}/{url}", **self.args)
        elif storage in ("ssh://", "scp://", "sftp://"):
            # We need to quote parameters to deal with special characters
            # https://bugs.python.org/issue18140
            user = urllib.parse.quote(self._provider["user"])
            host = urllib.parse.quote(self._provider["host"])
            url = urllib.parse.quote(url)
            # TODO: Remove int casting when https://github.com/airbytehq/airbyte/issues/4952 is addressed
            # TODO: The "port" field in spec.json must also be changed
            _port_value = self._provider.get("port", 22)
            try:
                port = int(_port_value)
            except ValueError as err:
                raise ValueError(f"{_port_value} is not a valid integer for the port") from err
            # Explicitly turn off ssh keys stored in ~/.ssh
            transport_params = {"connect_kwargs": {"look_for_keys": False}, "timeout": SSH_TIMEOUT}
            if "auth_ssh_key" in self._provider:
                private_key_path = self.create_temp_file("private_key", self._provider["auth_ssh_key"])
                transport_params["connect_kwargs"]["key_filename"] = private_key_path
            if "password" in self._provider:
                password = urllib.parse.quote(self._provider["password"])
                uri = f"{storage}{user}:{password}@{host}:{port}/{url}"
            else:
                uri = f"{storage}{user}@{host}:{port}/{url}"
            return smart_open.open(uri, transport_params=transport_params, **self.args)
        elif storage in ("https://", "http://"):
            transport_params = None
            if "user_agent" in self._provider and self._provider["user_agent"]:
                airbyte_version = environ.get("AIRBYTE_VERSION", "0.0")
                transport_params = {"headers": {"Accept-Encoding": "identity", "User-Agent": f"Airbyte/{airbyte_version}"}}
            logger.info(f"TransportParams: {transport_params}")
            return smart_open.open(self.full_url, transport_params=transport_params, **self.args)
        return smart_open.open(self.full_url, **self.args)

    @property
    def url(self) -> str:
        """Convert URL to remove the URL prefix (scheme)
        :return: the corresponding URL without URL prefix / scheme
        """
        parse_result = urlparse(self._url)
        if parse_result.scheme:
            return self._url.split("://")[-1]
        else:
            return self._url

    @property
    def storage_scheme(self) -> str:
        """Convert Storage Names to the proper URL Prefix
        :return: the corresponding URL prefix / scheme
        """
        storage_name = self._provider["storage"].upper()
        parse_result = urlparse(self._url)

        if storage_name == "GCS":
            return "gs://"
        elif storage_name == "S3":
            return "s3://"
        elif storage_name == "AZBLOB":
            return "azure://"
        elif storage_name == "HTTPS":
            return "https://"
        elif storage_name == "SSH" or storage_name == "SCP":
            return "scp://"
        elif storage_name == "SFTP":
            return "sftp://"
        elif storage_name == "WEBHDFS":
            return "webhdfs://"
        elif storage_name == "LOCAL":
            return "file://"
        elif parse_result.scheme:
            return parse_result.scheme

        logger.error(f"Unknown Storage provider in: {self._url}")
        return ""

    def _open_gcs_url(self) -> object:
        service_account_json = self._provider.get("service_account_json")
        credentials = None
        if service_account_json:
            try:
                credentials = json.loads(self._provider["service_account_json"])
            except json.decoder.JSONDecodeError as err:
                error_msg = f"Failed to parse gcs service account json: {repr(err)}"
                logger.error(f"{error_msg}\n{traceback.format_exc()}")
                raise ConfigurationError(error_msg) from err

        if credentials:
            credentials = service_account.Credentials.from_service_account_info(credentials)
            client = GCSClient(credentials=credentials, project=credentials._project_id)
        else:
            client = GCSClient.create_anonymous_client()
        file_to_close = smart_open.open(self.full_url, transport_params={"client": client}, **self.args)

        return file_to_close

    def _open_aws_url(self):
        aws_access_key_id = self._provider.get("aws_access_key_id")
        aws_secret_access_key = self._provider.get("aws_secret_access_key")
        use_aws_account = aws_access_key_id and aws_secret_access_key

        if use_aws_account:
            aws_access_key_id = self._provider.get("aws_access_key_id", "")
            aws_secret_access_key = self._provider.get("aws_secret_access_key", "")
            aws_region = self._provider.get("region", None)

            client = boto3.client(
                "s3", aws_access_key_id=aws_access_key_id, aws_secret_access_key=aws_secret_access_key, region_name=aws_region
            )
            result = smart_open.open(self._url, transport_params=dict(client=client), **self.args)
        else:
            config = botocore.client.Config(signature_version=botocore.UNSIGNED)
            params = {"client": boto3.client("s3", config=config)}
            result = smart_open.open(self.full_url, transport_params=params, **self.args)
        return result

    def _open_azblob_url(self):
        storage_account = self._provider.get("storage_account")
        storage_acc_url = f"https://{storage_account}.blob.core.windows.net"
        sas_token = self._provider.get("sas_token", None)
        shared_key = self._provider.get("shared_key", None)
        # if both keys are provided, shared_key is preferred as has permissions on entire storage account
        credential = shared_key or sas_token

        if credential:
            client = BlobServiceClient(account_url=storage_acc_url, credential=credential)
        else:
            # assuming anonymous public read access given no credential
            client = BlobServiceClient(account_url=storage_acc_url)

        url = f"{self.storage_scheme}{self.url}"
        return smart_open.open(url, transport_params=dict(client=client), **self.args)


class Client:
    """Class that manages reading and parsing data from streams"""

    CSV_CHUNK_SIZE = 10_000
    PARQUET_BATCH_SIZE = 1000
    reader_class = URLFile
    binary_formats = {"excel", "excel_binary", "feather", "parquet", "orc", "pickle"}

    def __init__(
            self, dataset_name: str, url: str, provider: dict, format: str = None, reader_options: dict = None,
            encryption_options: dict = None
    ):
        self._dataset_name = dataset_name
        self._url = url
        self._provider = provider
        self._reader_format = format or "csv"
        self._reader_options = reader_options or {}
        self.binary_source = self._reader_format in self.binary_formats or encryption_options
        self.encoding = self._reader_options.get("encoding")
        self.encryption_options = encryption_options

    @property
    def reader_class(self):
        if is_cloud_environment():
            return URLFileSecure

        return URLFile

    @property
    def stream_name(self) -> str:
        file = urlparse(self._url)
        if self._dataset_name:
            return f"file_{self._dataset_name}"
        return f"file_{self._provider['storage']}_{file.path.split('/')[-1]}"

    def load_nested_json_schema(self, fp) -> dict:
        # Use Genson Library to take JSON objects and generate schemas that describe them,
        builder = SchemaBuilder()
        for o in self.read():
            builder.add_object(o)

        result = builder.to_schema()
        if "items" in result:
            # this means we have a json list e.g. [{...}, {...}]
            # but need to emit schema of an inside dict
            result = result["items"]
        result["$schema"] = "http://json-schema.org/draft-07/schema#"
        return result

    def load_nested_json(self, fp) -> list:
        if self._reader_format == "jsonl":
            result = []
            line = fp.readline()
            while line:
                result.append(json.loads(line))
                line = fp.readline()
        else:
            result = json.load(fp)
            if not isinstance(result, list):
                result = [result]
        # for json and jsonl
        result = [{**d, '_ab_source_file_url': self._url} for d in result]
        return result

    def load_yaml(self, fp):
        if self._reader_format == "yaml":
            return pd.DataFrame(safe_load(fp))

    def load_dataframes(self, fp, skip_data=False, read_sample_chunk: bool = False) -> Iterable:
        """load and return the appropriate pandas dataframe.

        :param fp: file-like object to read from
        :param skip_data: limit reading data
        :param read_sample_chunk: indicates whether a single chunk should only be read to generate schema
        :return: a list of dataframe loaded from files described in the configuration
        """
        readers = {
            # pandas.read_csv additional arguments can be passed to customize how to parse csv.
            # see https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_csv.html
            "csv": pd.read_csv,
            # We can add option to call to pd.normalize_json to normalize semi-structured JSON data into a flat table
            # by asking user to specify how to flatten the nested columns
            "flat_json": pd.read_json,
            "html": pd.read_html,
            "excel": pd.read_excel,
            "excel_binary": pd.read_excel,
            "fwf": pd.read_fwf,
            "feather": pd.read_feather,
            "parquet": pq.ParquetFile,
            "orc": pd.read_orc,
            "pickle": pd.read_pickle,
        }

        try:
            reader = readers[self._reader_format]
        except KeyError as err:
            error_msg = f"Reader {self._reader_format} is not supported."
            logger.error(f"{error_msg}\n{traceback.format_exc()}")
            raise ConfigurationError(error_msg) from err

        reader_options = {**self._reader_options}

        try:
            if self._reader_format == "csv":
                bytes_read = 0
                reader_options["chunksize"] = self.CSV_CHUNK_SIZE
                if skip_data:
                    reader_options["nrows"] = 0
                    reader_options["index_col"] = 0
                for record in reader(fp, **reader_options):
                    bytes_read += sys.getsizeof(record)
                    yield record
                    if read_sample_chunk and bytes_read >= self.CSV_CHUNK_SIZE:
                        return
            elif self._reader_format == "excel_binary":
                reader_options["engine"] = "pyxlsb"
                yield reader(fp, **reader_options)
            elif self._reader_format == "excel":
                # Use openpyxl to read new-style Excel (xlsx) file; return to pandas for others
                try:
                    yield from self.openpyxl_chunk_reader(fp, **reader_options)
                except (InvalidFileException, BadZipFile):
                    yield reader(fp, **reader_options)
            elif self._reader_format == "parquet":
                if "batch_size" not in reader_options:
                    reader_options["batch_size"] = self.CSV_CHUNK_SIZE
                yield from reader(fp).iter_batches(batch_size=reader_options["batch_size"])
            else:
                yield reader(fp, **reader_options)
        except ParserError as err:
            error_msg = f"File {fp} can not be parsed. Please check your reader_options. https://pandas.pydata.org/pandas-docs/stable/user_guide/io.html"
            logger.error(f"{error_msg}\n{traceback.format_exc()}")
            raise AirbyteTracedException(message=error_msg, internal_message=error_msg, failure_type=FailureType.config_error) from err
        except UnicodeDecodeError as err:
            error_msg = (
                f"File {fp} can't be parsed with reader of chosen type ({self._reader_format}). "
                f"Please check provided Format and Reader Options. {repr(err)}."
            )
            logger.error(f"{error_msg}\n{traceback.format_exc()}")
            raise ConfigurationError(error_msg) from err

    @staticmethod
    def dtype_to_json_type(current_type: str, dtype) -> str:
        """Convert Pandas Dataframe types to Airbyte Types.

        :param current_type: str - one of the following types based on previous dataframes
        :param dtype: Pandas Dataframe type
        :return: Corresponding Airbyte Type
        """
        number_types = ("double", "float64")
        integer_types = ("int32", "int64", "int96")
        if current_type == "string":
            # previous column values was of the string type, no sense to look further
            return current_type
        if dtype == object:
            return "string"
        if str(dtype).lower() in number_types:
            return "number"
        if str(dtype).lower() in integer_types:
            return "integer" if not current_type else current_type
        if str(dtype).lower() == "bool" and (not current_type or current_type == "boolean"):
            return "boolean"
        if dtype == "datetime64[ns]":
            return "date-time"
        return "string"

    @staticmethod
    def dtype_to_json_type2(dtype):
        """Convert Pandas Dataframe types to Airbyte Types.

        :param dtype: Pandas Dataframe type
        :return: Corresponding Airbyte Type
        """
        number_types = ("double", "float64", "decimal128(10,2)", "decimal128(9,2)")
        integer_types = ("int32", "int64", "int96")
        datetime_types = ("datetime64[ns]", "timestamp[us]")
        decimal_pattern = r'decimal(\d+\(\d+,\d+\)|128\(\d+,\d+\))?'
        if dtype == object:
            return "string", None
        elif re.match(decimal_pattern, str(dtype)):
            return "number", "double"
        if str(dtype).lower() in integer_types:
            return "integer", None
        if str(dtype).lower() == "bool":
            return "boolean", None
        if dtype in datetime_types:
            return "date-time", None
        return "string", None

    @property
    def reader(self) -> reader_class:
        return self.reader_class(url=self._url, provider=self._provider, binary=self.binary_source, encoding=self.encoding)

    @backoff.on_exception(backoff.expo, ConnectionResetError, on_backoff=backoff_handler, max_tries=5, max_time=60)
    def read(self, fields: Iterable = None) -> Iterable[dict]:
        """Read data from the stream"""
        with self.reader.open() as fp:
            try:
                file_path = ""
                if self.encryption_options and self.encryption_options["encryption_method"] == "PGP":
                    file_path = f"/tmp/plain.{self._reader_format}"
                    Pgp(**self.encryption_options).decrypt(fp, file_path)
                    fp = s_open(file_path, 'r')
                if self._reader_format in ["json", "jsonl"]:
                    yield from self.load_nested_json(fp)
                elif self._reader_format == "yaml":
                    fields = set(fields) if fields else None
                    df = self.load_yaml(fp)
                    columns = fields.intersection(set(df.columns)) if fields else df.columns
                    df = df.where(pd.notnull(df), None)
                    yield from df[columns].to_dict(orient="records")
                else:
                    fields = frozenset(fields) if fields else None
                    # if self.binary_source:
                    #     fp = self._cache_stream(fp)
                    for batch in self.load_dataframes(fp):
                        df = batch.to_pandas() if self._reader_format == "parquet" else batch
                        # for parquet files
                        df['_ab_source_file_url'] = self._url
                        df_cols = list(df.columns)
                        columns = [x for x in df_cols if x in fields] if fields else df.columns
                        df.replace({np.nan: None}, inplace=True)
                        yield from df[list(columns)].to_dict(orient="records")
            except ConnectionResetError:
                logger.info(f"Catched `connection reset error - 104`, stream: {self.stream_name} ({self.reader.full_url})")
                raise ConnectionResetError
            except ProtocolError as err:
                error_msg = (
                    f"File {fp} can not be opened due to connection issues on provider side. Please check provided links and options"
                )
                logger.error(f"{error_msg}\n{traceback.format_exc()}")
                raise ConfigurationError(error_msg) from err
            except ParserError as err:
                error_msg = f"File {fp} can not be parsed. Please check your reader_options. https://pandas.pydata.org/pandas-docs/stable/user_guide/io.html"
                logger.error(f"{error_msg}\n{traceback.format_exc()}")
                raise ConfigurationError(error_msg) from err
            finally:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    logger.info(f"The file at {file_path} has been deleted.")

    def _cache_stream(self, fp):
        """cache stream to file"""
        fp_tmp = tempfile.TemporaryFile(mode="w+b")
        fp_tmp.write(fp.read())
        fp_tmp.seek(0)
        fp.close()
        return fp_tmp

    def _stream_properties(self, fp, empty_schema: bool = False, read_sample_chunk: bool = False):
        """
        empty_schema param is used to check connectivity, i.e. we only read a header and do not produce stream properties
        read_sample_chunk is used to determine if just one chunk should be read to generate schema
        """
        row_count = 0
        fields = {}
        if self._reader_format == "yaml":
            df_list = [self.load_yaml(fp)]

        elif self._reader_format == "parquet":
            parquet_file = pq.ParquetFile(fp)

            # Get the schema from the Parquet file
            schema = parquet_file.schema.to_arrow_schema()
            for name, dtype in zip(schema.names, schema.types):
                json_type, airbyte_type = self.dtype_to_json_type2(dtype)
                fields[name] = {"type": json_type}
                if airbyte_type:
                    fields[name]["airbyte_type"] = airbyte_type
            df_list = []
        else:
            # if self.binary_source:
            #     fp = self._cache_stream(fp)
            #     logger.info("Cache stream successs")
            df_list = self.load_dataframes(fp, skip_data=False)
        for df in df_list:
            df = df.to_pandas() if self._reader_format == "parquet" else df
            for col in df.columns:
                if df[col].isnull().values.all():
                    if not fields.get(col):
                        fields[col] = {"type": None}
                    continue
                # if data type of the same column differs in dataframes, we choose the broadest one
                prev_frame_column_type = fields.get(col, {}).get("type")
                fields[col] = {"type": self.dtype_to_json_type(prev_frame_column_type, df[col].dtype), "dtype": df[col].dtype}

                if is_timedelta(df[col]):
                    fields[col]["format"] = "date-time"
                    fields[col]["airbyte_type"] = "timestamp_with_timezone"
                elif is_datetime(df[col]):
                    fields[col]["format"] = "date-time"

        stream = {}
        fields['_ab_source_file_url'] = {'type': 'string'}
        for field in fields:
            stream[field] = {"type": [fields[field]["type"] if fields[field]["type"] else "string", "null"]}
            if "format" in fields[field]:
                stream[field]["format"] = fields[field]["format"]
            if "airbyte_type" in fields[field]:
                stream[field]["airbyte_type"] = fields[field]["airbyte_type"]
        return {"stream": stream, "row_count": row_count}

    def streams(self, empty_schema: bool = False) -> Iterable:
        """Discovers available streams"""
        # TODO handle discovery of directories of multiple files instead
        with self.reader.open() as fp:
            file_path = ""
            if self.encryption_options and self.encryption_options["encryption_method"] == "PGP":
                file_path = f"/tmp/plain.{self._reader_format}"
                Pgp(**self.encryption_options).decrypt(fp, file_path)
                fp = s_open(file_path, 'r')
            if self._reader_format in ["json", "jsonl"]:
                json_schema = self.load_nested_json_schema(fp)
            else:
                stream_property = self._stream_properties(fp)
                json_schema = {
                    "$schema": "http://json-schema.org/draft-07/schema#",
                    "type": "object",
                    "properties": stream_property["stream"],
                    "row_count": stream_property["row_count"],
                }
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"The file at {file_path} has been deleted.")
        yield AirbyteStream(name=self.stream_name, json_schema=json_schema, supported_sync_modes=[SyncMode.full_refresh])

    def openpyxl_chunk_reader(self, file, **kwargs):
        """Use openpyxl lazy loading feature to read excel files (xlsx only) in chunks of 500 lines at a time"""
        work_book = load_workbook(filename=file)
        user_provided_column_names = kwargs.get("names")
        for sheetname in work_book.sheetnames:
            work_sheet = work_book[sheetname]
            data = work_sheet.values
            end = work_sheet.max_row
            if end == 1 and not user_provided_column_names:
                message = "Please provide column names for table in reader options field"
                logger.error(message)
                raise AirbyteTracedException(
                    message="Config validation error: " + message,
                    internal_message=message,
                    failure_type=FailureType.config_error,
                )
            cols, start = (next(data), 1) if not user_provided_column_names else (user_provided_column_names, 0)
            step = 500
            while start <= end:
                df = pd.DataFrame(data=(next(data) for _ in range(start, min(start + step, end))), columns=cols)
                yield df
                start += step


class URLFileSecure(URLFile):
    """Updating of default logic:
    This connector shouldn't work with local files.
    """

    def __init__(self, url: str, provider: dict, binary=None, encoding=None):
        storage_name = provider["storage"].lower()
        if url.startswith("file://") or storage_name == LOCAL_STORAGE_NAME:
            raise RuntimeError("the local file storage is not supported by this connector.")
        super().__init__(url, provider, binary, encoding)
